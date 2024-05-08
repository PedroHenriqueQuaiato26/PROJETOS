import tkinter as tk
from tkinter import filedialog
from functools import partial
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os
from openpyxl import load_workbook
import time 
from flask import Flask, render_template, request
import re

#-------------------------------------------------------------------------------------------------------------------------#

app = Flask(__name__)

#-------------------------------------------------------------------------------------------------------------------------#

def escrever_pdf(nome_pdf, mensagem):
    try:
        with open(nome_pdf, 'w', encoding='utf-8') as pdf_file:
            pdf_file.write(mensagem)
    except Exception as e:
        print(f"Erro ao escrever PDF {nome_pdf}: {e}")

#-------------------------------------------------------------------------------------------------------------------------#

def substituir_caractere_excel(caminho_arquivo_excel, caractere_a_substituir):
    workbook = load_workbook(filename=caminho_arquivo_excel)
    sheet = workbook.active
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.replace(caractere_a_substituir, '-')
                
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.replace(caractere_a_substituir2, '.')
    
    workbook.save(caminho_arquivo_excel)
    
    global arquivo_carregado
    arquivo_carregado = pd.read_excel(caminho_arquivo_excel)

caractere_a_substituir = ';'
caractere_a_substituir2 = ','

#-------------------------------------------------------------------------------------------------------------------------#

def analisar_identificacao(arquivo_carregado,nome_pdf, pasta_relatorios=None):
    try:
        identificacao = arquivo_carregado['Identificacao']
        if (identificacao == 'PR').all():
            mensagem_acerto = 'Todos os elementos são PR'
            return mensagem_acerto, None
        else:
            arquivo_carregado.loc[arquivo_carregado['Identificacao'] != 'PR', 'Identificacao'] = 'PR'
            mensagem_acerto = 'Alterações realizadas com sucesso!'
            arquivo_carregado.to_excel(os.path.join(pasta_relatorios, 'Modelo_Pro.xlsx'), index=False)
            return mensagem_acerto, None
    except Exception as e:
        mensagem_erro = f"Ocorreu um erro na análise de identificação: {e}"
        return None, mensagem_erro

#-------------------------------------------------------------------------------------------------------------------------#


def analisar_codigo(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        codigo = arquivo_carregado['codigo']
        if (codigo == arquivo_carregado['codigo'].astype(int)).all():
            mensagem_acerto = 'Os dados da Coluna codigo são do tipo inteiro'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        elif codigo.isnull().all():
            mensagem_acerto = 'Todos os códigos são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna!'
            escrever_pdf(os.path.join(pasta_relatorios, 'Relatorio_de_Erros.pdf'), mensagem_erro)
    except Exception as e:
        mensagem_erro = f"Ocorreu um erro na análise de código: {e}"
        escrever_pdf(os.path.join(pasta_relatorios, 'Erro_ao_Processar_Arquivo_codigo.pdf'), mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#



def analisar_referencia(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        referencia = arquivo_carregado['Referencia']
        if (referencia == arquivo_carregado['Referencia'].astype(str)).all():
            mensagem_acerto = 'Todas as referências são válidas, pois são Strings'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        elif (arquivo_carregado['Referencia'] == arquivo_carregado['Referencia'].replace(';', '-')).all():
            mensagem_acerto = 'Trocamos os sinais de ; por -'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        elif arquivo_carregado['Referencia'].isnull().any():
            mensagem_acerto = 'Existem valores nulos na coluna, mas eles podem ser válidos também'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        elif referencia.isnull().all():
            mensagem_acerto = 'Todos os códigos são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são Strings na coluna!'
            escrever_pdf(os.path.join(pasta_relatorios, 'Relatorio_de_Erros_Referencia.pdf'), mensagem_erro)
    except Exception as e:
        mensagem_erro = f"Ocorreu um erro na análise de referência: {e}"
        escrever_pdf(os.path.join(pasta_relatorios, 'Erro_ao_Processar_Arquivo_Referencia.pdf'), mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def analisar_codigo_ean(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []
        if not arquivo_carregado['codigo_ean'].astype(str).str.isalnum().all():
            mensagens_erro.append('Código EAN inválido, pois contém caracteres especiais.')
        else:
            mensagens_acerto.append('Código EAN válido, não contém caracteres especiais.')
        if arquivo_carregado['codigo_ean'].isnull().any():
            mensagens_acerto.append('Código EAN valido, pois contém valores nulos.')
        else:
            mensagens_acerto.append('Código EAN válido, não contém valores nulos.')
        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            escrever_pdf(os.path.join(pasta_relatorios, 'Relatorio_de_Erros_codigo_ean.pdf'), mensagem_erro)
        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
    except Exception as e:
        mensagem_erro = f"Ocorreu um erro na análise de código EAN: {e}"
        escrever_pdf(os.path.join(pasta_relatorios, 'Erro_ao_Processar_Arquivo_codigo_ean.pdf'), mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def analisar_nome(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        if arquivo_carregado['Nome'].str.contains(';').any():
            mensagens_erro.append('Coluna Nome inválida, pois alguns dados possuem ponto e vírgula')
        else:
            mensagens_acerto.append('A Coluna Nome é válida')

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            escrever_pdf(os.path.join(pasta_relatorios, 'Relatorio_de_Erros_nome.pdf'), mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        escrever_pdf(os.path.join(pasta_relatorios, 'Erro_ao_Processar_Arquivo_nome.pdf'), mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def Tipo(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if not (arquivo_carregado['Tipo'].astype(str).isin(['P', 'S'])).all():
            mensagens_erro.append('Coluna Tipo inválida, pois os dados não são iguais a P ou S')
        else:
            mensagens_acerto.append('Coluna Tipo válida, pois possui os dados P e S')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Tipo.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Tipo.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)
#-------------------------------------------------------------------------------------------------------------------------#


def Codigo_do_Fornecedor(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        coluna_codigo_fornecedor = arquivo_carregado['Codigo_do_Fornecedor']

        if coluna_codigo_fornecedor.dtype in ['int', 'int64','float','float64']:
            mensagem_acerto = 'Todos os códigos são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif coluna_codigo_fornecedor.isnull().all():
            mensagem_acerto = 'Todos os códigos são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            # Verifica quais valores específicos não são numéricos
            valores_nao_numericos = coluna_codigo_fornecedor[~coluna_codigo_fornecedor.astype(str).str.isdigit()]
            mensagem_erro = f'Alguns valores na coluna Código do fornecedor não são numéricos: {valores_nao_numericos.tolist()}'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Codigo_do_Fornecedor.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Codigo_do_Fornecedor.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def Preco(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    try:
        preco = arquivo_carregado['Preco']

        if arquivo_carregado['Preco'].astype(float).all():
            mensagem_acerto = 'Todos os códigos são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        elif preco.equals(preco.astype(int)):
            mensagem_acerto = 'Todos os códigos são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna!'
            escrever_pdf(os.path.join(pasta_relatorios, 'Relatorio_de_Erros_Preco.pdf'), mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Preco.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def unidade_de_medida(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if not (arquivo_carregado['unidade_de_medida'].astype(str).isin(['AP', 'BL', 'CJ', 'CX', 'DZ', 'FD', 'GL', 'GR',
                                                                         'JR', 'KG', 'LA', 'LT', 'M2', 'M3', 'MI', 'MT',
                                                                         'PC', 'PR', 'PT', 'RL', 'TQ', 'UN', 'CN', 'HR'])).all():
            mensagens_erro.append('Coluna Pode conter erros, porém se a unidade de medida que estiver tentando '
                                  'colocar não for validada, será preciso cadastrar no sistema, pois não impedirá a '
                                  'importação')
        else:
            mensagens_acerto.append('Coluna Unidade de medida válida pois possui unidades que são nativas do sistema')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_unidade_de_medida.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_unidade_de_medida.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def Lucro(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Lucro = arquivo_carregado['Lucro']

        if Lucro.dtype == 'int64':
            mensagem_acerto = 'Todos os Lucro são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif Lucro.dtype == 'int':
            mensagem_acerto = 'Todos os Lucro são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
            
        elif Lucro.dtype == 'float':
            mensagem_acerto = 'Todos os Lucro são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif Lucro.dtype == 'float64':
            mensagem_acerto = 'Todos os Lucro são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif Lucro.isnull().all():
            mensagem_acerto = 'Todos os Lucro são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna Código do fornecedor!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Lucro.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Lucro.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#-------------------------------------------------------------------------------------------------------------------------#


def Peso(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Peso = arquivo_carregado['Peso']

        if Peso.dtype == 'float' or Peso.dtype == 'int64':
            mensagem_acerto = 'Todos os códigos são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        elif Peso.isnull().all():
            mensagem_acerto = 'Todos os Pesos são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Peso.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Peso.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def Numero_de_serie(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Numero_de_serie = arquivo_carregado['Numero_de_serie']

        if Numero_de_serie.dtype == 'int64':
            mensagem_acerto = 'Todos os Numero_de_serie são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif Numero_de_serie.dtype == 'int':
            mensagem_acerto = 'Todos os Numero_de_serie são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
            
        
        elif Numero_de_serie.isnull().all():
            mensagem_acerto = 'Todos os Numero_de_serie são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna Numero_de_serie!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Numero_de_serie.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Numero_de_serie.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#-------------------------------------------------------------------------------------------------------------------------#


def Tributação_ICMS(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if (arquivo_carregado['Tributação_ICMS'].astype(str).isin(['T','N','I','F','S','NS','IS','FS'])).all():
            mensagens_acerto.append('Unidades de Tributação estão corretas e de '
                                  'com a ferramenta que verifica as importações ')
        elif arquivo_carregado['Tributação_ICMS'].isnull().all():
            mensagem_acerto = 'Todos os códigos são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagens_acerto.append('Coluna Tributação ICMS válida pois possui unidades que são nativas do sistema')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Tributação_ICMS.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Tributação_ICMS.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#-------------------------------------------------------------------------------------------------------------------------#


def IPI(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        IPI = arquivo_carregado['IPI']

        if IPI.dtype == 'int64':
            mensagem_acerto = 'Todos os IPI são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif IPI.dtype == 'int':
            mensagem_acerto = 'Todos os IPI são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
            
        elif IPI.dtype == 'float':
            mensagem_acerto = 'Todos os IPI são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif IPI.dtype == 'float64':
            mensagem_acerto = 'Todos os IPI são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif IPI.isnull().all():
            mensagem_acerto = 'Todos os IPI são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna Código do fornecedor!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_IPI.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_IPI.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def situacao_tributaria(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        situacao_tributaria = arquivo_carregado['situacao_tributaria']

        if situacao_tributaria.dtype == 'int64':
            mensagem_acerto = 'Todos os situacao_tributaria são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif situacao_tributaria.dtype == 'int':
            mensagem_acerto = 'Todos ossituacao_tributaria são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
            
        
        elif situacao_tributaria.isnull().all():
            mensagem_acerto = 'Todos os situacao_tributaria são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna situacao_tributaria!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_situacao_tributaria.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_situacao_tributaria.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)



#-------------------------------------------------------------------------------------------------------------------------#


def Custo(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Custo = arquivo_carregado['Custo']

        if Custo.dtype == 'float64' or Custo.dtype == 'float' or Custo.dtype == 'int64':
            mensagem_acerto = 'Todos os códigos são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif Custo['Custo'].isnull().all():
            mensagem_acerto = 'Todos os códigos são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna Custo!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Peso.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Custo")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def IAT(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if (arquivo_carregado['IAT'].astype(str).isin(['A','T'])).all():
            mensagens_acerto.append('IAT estão corretas ')
        elif arquivo_carregado['IAT'].isnull().all():
            mensagem_acerto = 'Todos os IAT são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagens_erro.append('Coluna IAT inválida pois não possui unidades que são nativas do sistema')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_IAT.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_IAT.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def IPPT(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if (arquivo_carregado['IPPT'].astype(str).isin(['P','T'])).all():
            mensagens_acerto.append('IPPT estão no padrão')
        elif arquivo_carregado['IPPT'].isnull().all():
            mensagem_acerto = 'Todos os IPPT são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagens_acerto.append('Coluna IPPT válida pois possui unidades que são nativas do sistema')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_IPPT.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_IPPT.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def Origem(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if (arquivo_carregado['Origem'].astype(str).isin(['0', '1'])).all():
            mensagens_acerto.append('Coluna Tipo válida, pois os dados são iguais a 0 ou 1')
        elif arquivo_carregado['Origem'].isnull().all():
            mensagem_acerto = 'Todos os Origem são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagens_acerto.append('Coluna Tipo válida, pois possui os dados 0 e 1')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Origem.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Origem.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def Grupo(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        if arquivo_carregado['Grupo'].str.contains(';').all():
            mensagens_erro.append('Coluna Grupo inválida, pois alguns dados possuem ponto e vírgula')
        else:
            mensagens_acerto.append('A Coluna Grupo é válida')

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            escrever_pdf(os.path.join(pasta_relatorios, 'Relatorio_de_Erros_Grupo.pdf'), mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        escrever_pdf(os.path.join(pasta_relatorios, 'Erro_ao_Processar_Arquivo_Grupo.pdf'), mensagem_erro)
       

#-------------------------------------------------------------------------------------------------------------------------#


def Fornecedor(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Fornecedor = ['Fornecedor']

        for coluna in Fornecedor:
            if arquivo_carregado[coluna].dtype == 'object' and arquivo_carregado[coluna].str.contains(';').any():
                mensagens_erro.append(f'Coluna {coluna} inválida, pois alguns dados possuem ponto e vírgula')
        
        if not mensagens_erro:
            mensagem_acerto = 'Todos os dados são válidos, nenhum ponto e vírgula encontrado.'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Fornecedor.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Fornecedor.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def Caminho_da_imagem(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        if arquivo_carregado['Caminho_da_imagem'].astype(str).str.contains(';').all():
            mensagens_erro.append('Coluna Grupo inválida, pois alguns dados possuem ponto e vírgula')
        
        elif arquivo_carregado['Caminho_da_imagem'].astype(str).str.contains("'\'").all():
            mensagens_acerto.append('Todas as URL são validas, porem verifique como se todas estão'
                                    'presentes em sua maquina local ou se imagens da web são validas')
        
        elif arquivo_carregado['Caminho_da_imagem'].isnull().all():
            mensagem_acerto = 'Todos os Caminho da imagem são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagens_acerto.append('A Coluna Caminho da imagem é válida')

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            escrever_pdf(os.path.join(pasta_relatorios, 'Relatorio_de_Erros_Caminho_da_imagem.pdf'), mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        escrever_pdf(os.path.join(pasta_relatorios, 'Erro_ao_Processar_Arquivo_Caminho_da_imagem.pdf'), mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def ICMS(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        ICMS = arquivo_carregado['ICMS']

        if ICMS.dtype == 'int64' or ICMS.dtype == 'int' or ICMS.dtype == 'float' or ICMS.dtype == 'float64':
            mensagem_acerto = 'Todos os códigos são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif arquivo_carregado['ICMS'].isnull().all():
            mensagem_acerto = 'Todos os ICMS são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna ICMS!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_ICMS.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_ICMS")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#-------------------------------------------------------------------------------------------------------------------------#


def Tributacao_especial(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Tributacao_especial = arquivo_carregado['Tributacao_especial']

        if Tributacao_especial.dtype == 'int64' or Tributacao_especial.dtype == 'int':
            mensagem_acerto = 'Todos os Tributacao especial são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif arquivo_carregado['Tributacao_especial'].isnull().all():
            mensagem_acerto = 'Todos os Tributacao Especial são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna Tributacao Especial!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Tributacao_especial.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Tributacao_especial")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def Casas_decimais_da_unidade_de_medida(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Casas_decimais_da_unidade_de_medida = arquivo_carregado['Casas_decimais_da_unidade_de_medida']

        if Casas_decimais_da_unidade_de_medida.dtype == 'int64' or Casas_decimais_da_unidade_de_medida.dtype == 'int':
            mensagem_acerto = 'Todos os Casas_decimais_da_unidade_de_medida são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        elif Casas_decimais_da_unidade_de_medida.isnull().all():
            mensagem_acerto = 'Todos os Casas_decimais_da_unidade_de_medida são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna Casas_decimais_da_unidade_de_medida!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Casas_decimais_da_unidade_de_medida.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Casas_decimais_da_unidade_de_medida")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def codigo_de_grupo(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        codigo_de_grupo = arquivo_carregado['codigo_de_grupo']

        if  codigo_de_grupo.dtype == 'int64':
            mensagem_acerto = 'Todos os codigo_de_grupo são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif arquivo_carregado['codigo_de_grupo'].isnull().all():
            mensagem_acerto = 'Todos os codigo_de_grupo são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna codigo_de_grupo!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_codigo_de_grupo.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_codigo_de_grupo")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#-------------------------------------------------------------------------------------------------------------------------#


def Pesavel(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if (arquivo_carregado['Pesavel'].astype(str).isin(['0', '1'])).all():
            mensagens_acerto.append('Coluna Pesavel válida, pois os dados são iguais a 0 ou 1')
        elif arquivo_carregado['Pesavel'].isnull().all():
            mensagem_acerto = 'Todos os Pesavel são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagens_acerto.append('Coluna Pesavel válida, pois possui os dados 0 e 1')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Pesavel.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Pesavel.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#-------------------------------------------------------------------------------------------------------------------------#


def Tipo_de_produto(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if not (arquivo_carregado['tipo_produto'].astype(str).isin(['00', '01','02','03','04','05','06','07','08','09','10','99'])).all():
            mensagens_erro.append('Coluna tipo_produto inválida, pois os dados não são iguais oas padrões verifique se os numeros tem zero')
        else:
            mensagens_acerto.append('Coluna tipo_produto válida, pois possui os dados 0 e 1')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_tipo_produto.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_tipo_produto.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def OBS(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        OBS_coluna_str = arquivo_carregado['OBS'].astype(str)

        if OBS_coluna_str.str.contains(';').any():
            mensagens_erro.append('Coluna OBS inválida, pois alguns dados possuem ponto e vírgula')
        else:
            mensagens_acerto.append('A Coluna OBS é válida')

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            escrever_pdf(os.path.join(pasta_relatorios, 'Relatorio_de_Erros_OBS.pdf'), mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        escrever_pdf(os.path.join(pasta_relatorios, 'Erro_ao_Processar_Arquivo_OBS.pdf'), mensagem_erro)
#-------------------------------------------------------------------------------------------------------------------------#


def Pautas_de_preco(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Pauta_preco1 = arquivo_carregado['Pauta_preco1']
        Pauta_preco2 = arquivo_carregado['Pauta_preco2']
        Pauta_preco3 = arquivo_carregado['Pauta_preco3']
        Pauta_preco4 = arquivo_carregado['Pauta_preco4']

        if  Pauta_preco1.dtype == 'int64' or 'float' and Pauta_preco2.dtype == 'int64' or 'float' and Pauta_preco3.dtype == 'int64' or 'float' and Pauta_preco4.dtype == 'int64' or 'float':
            mensagem_acerto = 'Todas as Pautas de Preço estão corretas são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif arquivo_carregado['Pauta_preco1'].isnull().all() and arquivo_carregado['Pauta_preco2'].isnull().all() and arquivo_carregado['Pauta_preco3'].isnull().all() and arquivo_carregado['Pauta_preco4'].isnull().all():
            mensagem_acerto = 'Todos os Pauta de Preço são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna Pautas_de_preco!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Pautas_de_preco.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Pautas_de_preco")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#-------------------------------------------------------------------------------------------------------------------------#

def NCM(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        NCM = arquivo_carregado['NCM']
        padrao_NCM = r'^\d{4}\.\d{2}\.\d{2}$'
        padrao_compilado = re.compile(padrao_NCM)

        if NCM.dtype == 'int64' or NCM.dtype == 'int' or NCM.dtype == 'float' or NCM.dtype == 'float64':
            mensagem_acerto = 'Todos os NCM são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif NCM.astype(str).str.match(padrao_compilado).all():
            mensagem_acerto = 'Todos os NCM estão válidos, pois estão dentro dos padrões'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif arquivo_carregado['NCM'].isnull().all():
            mensagem_acerto = 'Todos os NCM são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não estão dentro dos padrões na coluna NCM!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_NCM.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_NCM.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def analise(arquivo_carregado):
    mensagens_erro = []
    mensagens_acerto = []

    if not (arquivo_carregado['Tributacao_do_simples_nacional_nfe'].astype(str).isin(['101','102','103','201','202','203','300','400','100','900'])).all():
        mensagens_erro.append('Coluna Tributacao_do_simples_nacional_nfe inválida, pois os dados não são iguais às tributações')
    else:
        mensagens_acerto.append('Coluna Tributacao_do_simples_nacional_nfe válida, pois os dados são iguais às tributações')

    return mensagens_erro, mensagens_acerto

def Tributacao_do_simples_nacional_nfe(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro, mensagens_acerto = analise(arquivo_carregado) 

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Tributacao_do_simples_nacional_nfe.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Tributacao_do_simples_nacional_nfe.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)
#-------------------------------------------------------------------------------------------------------------------------#


def CST_Pis_Cofins_saida(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if (arquivo_carregado['CST_Pis_Cofins_saida'].astype(str).isin(['01','02','03','04','01','06','07','08','09','49','99'])).all():
            mensagens_erro.append('Coluna CST_Pis_Cofins_saida válida, pois os dados são iguais as bases nacionais')
        elif arquivo_carregado['CST_Pis_Cofins_saida'].isnull().all():
            mensagem_acerto = 'Todos os CST_Pis_Cofins_saida são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagens_acerto.append('Coluna CST_Pis_Cofins_saida válida, pois possui os dados 0 e 1')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CST_Pis_Cofins_saida.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CST_Pis_Cofins_saida.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def aliquota_pis_saida(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        aliquota_pis_saida = arquivo_carregado['aliquota_pis_saida']

        if aliquota_pis_saida.dtype == 'int64' or aliquota_pis_saida.dtype == 'int' or aliquota_pis_saida.dtype == 'float':
            mensagem_acerto = 'Todos os aliquota_pis_saida são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif arquivo_carregado['aliquota_pis_saida'].isnull().all():
            mensagem_acerto = 'Todos os aliquota_pis_saida são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna aliquota_pis_saida!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_aliquota_pis_saida.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_aliquota_pis_saida")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#-------------------------------------------------------------------------------------------------------------------------#


def Aliquota_confis_saida(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Aliquota_confis_saida = arquivo_carregado['Aliquota_confis_saida']

        if Aliquota_confis_saida.dtype == 'int64' or Aliquota_confis_saida.dtype == 'int' or Aliquota_confis_saida.dtype == 'float':
            mensagem_acerto = 'Todos os Aliquota_confis_saida são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif arquivo_carregado['Aliquota_confis_saida'].isnull().all():
            mensagem_acerto = 'Todos os Aliquota_confis_saida são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna Aliquota_confis_saida!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Aliquota_confis_saida.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Aliquota_confis_saida")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def CST_Pis_Cofins_entrada(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if (arquivo_carregado['CST_Pis_Cofins_entrada'].astype(str).isin(['10','11','12','13','14','1','16','60','61','62','63','64','61','66','67','70','71','72','73','74','71','98','99'])).all():
            mensagens_acerto.append('Coluna CST_Pis_Cofins_entrada válida, pois os dados são iguais ao Padrão nacional')
        elif arquivo_carregado['CST_Pis_Cofins_entrada'].isnull().all():
            mensagem_acerto = 'Todos os CST_Pis_Cofins_entrada são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagens_acerto.append('ColunaCST_Pis_Cofins_entrada válida, pois os dados são iguais ao Padrão nacional')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CST_Pis_Cofins_entrada.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CST_Pis_Cofins_entrada.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#-------------------------------------------------------------------------------------------------------------------------#


def aliquota_pis_entrada(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        aliquota_pis_entrada = arquivo_carregado['aliquota_pis_entrada']

        if aliquota_pis_entrada.dtype == 'int64' or aliquota_pis_entrada.dtype == 'int' or aliquota_pis_entrada.dtype == 'float':
            mensagem_acerto = 'Todos os aliquota_pis_entrada são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif arquivo_carregado['aliquota_pis_entrada'].isnull().all():
            mensagem_acerto = 'Todos os aliquota_pis_entrada são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna aliquota_pis_entrada!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_aliquota_pis_entrada.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_aliquota_pis_entrada")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def aliquota_cofins_entrada(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        aliquota_cofins_entrada = arquivo_carregado['aliquota_cofins_entrada']

        if aliquota_cofins_entrada.dtype == 'int64' or aliquota_cofins_entrada.dtype == 'int' or aliquota_cofins_entrada.dtype == 'float':
            mensagem_acerto = 'Todos os aliquota_cofins_entrada são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif arquivo_carregado['aliquota_cofins_entrada'].isnull().all():
            mensagem_acerto = 'Todos os aliquota_cofins_entrada são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna aliquota_cofins_entrada!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_aliquota_cofins_entrada.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_aliquota_cofins_entrada")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def Permite_informar_dimensoes(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if (arquivo_carregado['Permite_informar_dimensoes'].astype(str).isin(['0', '1'])).all():
            mensagens_acerto.append('Coluna Permite_informar_dimensoes válida, pois os dados são iguais a 0 ou 1')
        elif arquivo_carregado['Permite_informar_dimensoes'].isnull().all():
            mensagem_acerto = 'Todos os Permite_informar_dimensoes são válidos, pois podem ser Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagens_acerto.append('Coluna Permite_informar_dimensoes válida, pois possui os dados 0 e 1')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Permite_informar_dimensoes.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Permite_informar_dimensoes.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#------------------------------------------------------------------------------------------------------------------------#


def CFOP_interna_de_entrada(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_interna_de_entrada'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_interna_de_entrada'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_interna_de_entrada!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_saida.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_entrada.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_interna_de_entrada.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def CFOP_interna_de_saida(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_interna_de_saida'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_interna_de_saida'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_interna_de_saida!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_saida.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_saida.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_interna_de_saida.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def CFOP_externa_de_entrada(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_externa_de_entrada'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_externa_de_entrada'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_externa_de_entrada!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_entrada.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_entrada.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_externa_de_entrada.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#------------------------------------------------------------------------------------------------------------------------#


def CFOP_externa_de_saida(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_externa_de_saida'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_externa_de_saida'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_externa_de_saida!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_saida.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_saida.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_externa_de_saida.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#------------------------------------------------------------------------------------------------------------------------#


def CFOP_interna_de_entrada_devolucao(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_interna_de_entrada_devolucao'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_interna_de_entrada_devolucao'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_interna_de_entrada_devolucao!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_entrada_devolucao.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_entrada_devolucao.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_interna_de_entrada_devolucao.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def CFOP_interna_de_saida_devolucao(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_interna_de_saida_devolucao'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_interna_de_saida_devolucao'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_interna_de_saida_devolucao!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_saida_devolucao.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_saida_devolucao.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_interna_de_saida_devolucao.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def CFOP_externa_de_entrada_devolucao(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_externa_de_entrada_devolucao'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_externa_de_entrada_devolucao'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_externa_de_entrada_devolucao!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_entrada_devolucao.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_entrada_devolucao.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_externa_de_entrada_devolucao.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def CFOP_externa_de_saida_devolucao(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_externa_de_saida_devolucao'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_externa_de_saida_devolucao'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_externa_de_saida_devolucao!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_saida_devolucao.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_saida_devolucao.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_externa_de_saida_devolucao.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def CFOP_interna_de_entrada_transferencia(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_interna_de_entrada_transferencia'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_interna_de_entrada_transferencia'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_interna_de_entrada_transferencia!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_entrada_transferencia.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_entrada_transferencia.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_interna_de_entrada_transferencia.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def CFOP_interna_de_saida_transferencia(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_interna_de_saida_transferencia'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_interna_de_saida_transferencia'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_interna_de_saida_transferencia!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_saida_transferencia.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_interna_de_saida_transferencia.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_interna_de_saida_transferencia.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def CFOP_externa_de_entrada_transferencia(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_externa_de_entrada_transferencia'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_externa_de_entrada_transferencia'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_externa_de_entrada_transferencia!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_entrada_transferencia.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_entrada_transferencia.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_externa_de_entrada_transferencia.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def CFOP_externa_de_saida_transferencia(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if arquivo_carregado['CFOP_externa_de_saida_transferencia'].astype(str).str.match(r'^\d{4}$').all():
            mensagens_acerto.append('Todos os dados estão com 4 caracteres o que equivale ao numero da CFOP')
        elif arquivo_carregado['CFOP_externa_de_saida_transferencia'].isnull().all():
            mensagem_acerto = 'Todos os CFOP são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna CFOP_externa_de_saida_transferencia!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_saida_transferencia.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_CFOP_externa_de_saida_transferencia.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_CFOP_externa_de_saida_transferencia.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


import os

def informacao_extra(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        colunas_informacao_extra = ['Informacao_extra 1', 'Informacao_extra 2', 'Informacao_extra 3', 
                                    'Informacao_extra 4', 'Informacao_extra 5', 'Informacao_extra 6']

        for coluna in colunas_informacao_extra:
            if arquivo_carregado[coluna].dtype == 'object' and arquivo_carregado[coluna].str.contains(';').any():
                mensagens_erro.append(f'Coluna {coluna} inválida, pois alguns dados possuem ponto e vírgula')
        
        if not mensagens_erro:
            mensagem_acerto = 'Todos os dados são válidos, nenhum ponto e vírgula encontrado.'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Informacao_Extra.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Informacao_Extra.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#------------------------------------------------------------------------------------------------------------------------#


def cest(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []
        
        if arquivo_carregado['CEST'].astype(str).str.match(r'^\d{7}$').all():
            mensagens_acerto.append('Todos os dados estão com 7 caracteres o que equivale ao numero da cest')
        elif arquivo_carregado['CEST'].isnull().all():
            mensagem_acerto = 'Todos os cest são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna cest!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_cest.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_cest")
        escrever_pdf(nome_pdf_erro, mensagem_erro)
            

#------------------------------------------------------------------------------------------------------------------------#


def Informacao_adicional(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        # Convertendo os valores da coluna 'Informacao_adicional' para strings
        informacao_adicional_coluna = arquivo_carregado['Informacao_adicional'].astype(str)

        if informacao_adicional_coluna.str.contains(';').any():
            mensagens_erro.append('Coluna Informacao_adicional inválida, pois alguns dados possuem ponto e vírgula')
        else:
            mensagens_acerto.append('A Coluna Informacao_adicional é válida')

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            escrever_pdf(os.path.join(pasta_relatorios, 'Relatorio_de_Erros_Informacao_adicional.pdf'), mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        escrever_pdf(os.path.join(pasta_relatorios, 'Erro_ao_Processar_Arquivo_Informacao_adicional.pdf'), mensagem_erro)


#------------------------------------------------------------------------------------------------------------------------#

def Tributacao_do_Simples_Nacional_NFCe_ou_SAT(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    mensagens_erro = []
    mensagens_acerto = []

    def analise(arquivo_carregado):
        if (arquivo_carregado['Tributacao_do_Simples_Nacional_NFCe_ou_SAT'].astype(str).isin(['101', '102','103','201','202','203','300','400','100','900'])).all():
            mensagens_acerto.append('Coluna Tributacao_do_Simples_Nacional_NFCe_ou_SAT válida, pois os dados são iguais as tributações')
        elif arquivo_carregado['Tributacao_do_Simples_Nacional_NFCe_ou_SAT'].isnull().all():
            mensagem_acerto = 'Todos os códigos são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        else:
            mensagens_erro.append('Coluna Tributacao_do_Simples_Nacional_NFCe_ou_SAT inválida,' 
                                    'pois não possui os dados corretos em relação a Tributacao_do_Simples_Nacional_NFCe_ou_SAT')

    try:
        analise(arquivo_carregado)

        if mensagens_erro:
            mensagem_erro = '\n'.join(mensagens_erro)
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Tributacao_do_Simples_Nacional_NFCe_ou_SAT.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

        if mensagens_acerto:
            mensagem_acerto = '\n'.join(mensagens_acerto)
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Tributacao_do_Simples_Nacional_NFCe_ou_SAT.pdf")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def Custo_medio_inicial(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Custo_medio_inicial = arquivo_carregado['Custo_medio_inicial']

        if Custo_medio_inicial.dtype == 'float64' or Custo_medio_inicial.dtype == 'float' or Custo_medio_inicial.dtype == 'int64':
            mensagem_acerto = 'Todos os Custo_medio_inicial são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif Custo_medio_inicial['Custo_medio_inicial'].isnull().all():
            mensagem_acerto = 'Todos os Custo_medio_inicial são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna Custo_medio_inicial!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Custo_medio_inicial.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Custo_medio_inicial")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def Codigo_da_Lei_complementar(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Codigo_da_Lei_complementar = arquivo_carregado['Codigo_da_Lei_complementar']

        if arquivo_carregado['Custo_medio_inicial'].dtype == 'float64' or arquivo_carregado['Codigo_da_Lei_complementar'].dtype in ['float64', 'float', 'int64']:
            mensagem_acerto = 'Todos os Codigo_da_Lei_complementar são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif Codigo_da_Lei_complementar['Codigo_da_Lei_complementar'].isnull().all():
            mensagem_acerto = 'Todos os Custo_medio_inicial são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna Codigo_da_Lei_complementar!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Codigo_da_Lei_complementar.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Codigo_da_Lei_complementar")
        escrever_pdf(nome_pdf_erro, mensagem_erro)

#------------------------------------------------------------------------------------------------------------------------#


def Indicador_da_exigibilidade_do_ISS(arquivo_carregado, nome_pdf, pasta_relatorios=None):
    try:
        mensagens_erro = []
        mensagens_acerto = []

        Indicador_da_exigibilidade_do_ISS = arquivo_carregado['Indicador_da_exigibilidade_do_ISS']

        if Indicador_da_exigibilidade_do_ISS.dtype == 'float64' or Indicador_da_exigibilidade_do_ISS.dtype == 'float' or Indicador_da_exigibilidade_do_ISS.dtype == 'int64':
            mensagem_acerto = 'Todos os Codigo_da_Lei_complementar são válidos, pois são numéricos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)
        
        elif Indicador_da_exigibilidade_do_ISS['Indicador_da_exigibilidade_do_ISS'].isnull().all():
            mensagem_acerto = 'Todos os Custo_medio_inicial são válidos, pois são Nulos'
            escrever_pdf(os.path.join(pasta_relatorios, nome_pdf), mensagem_acerto)

        else:
            mensagem_erro = 'Temos dados que não são numéricos na coluna Indicador_da_exigibilidade_do_ISS!'
            nome_pdf_erro = os.path.join(pasta_relatorios, "Relatorio_de_Erros_Indicador_da_exigibilidade_do_ISS.pdf")
            escrever_pdf(nome_pdf_erro, mensagem_erro)

    except Exception as e:
        mensagem_erro = f"Ocorreu um erro: {e}"
        nome_pdf_erro = os.path.join(pasta_relatorios, "Erro_ao_Processar_Arquivo_Indicador_da_exigibilidade_do_ISS")
        escrever_pdf(nome_pdf_erro, mensagem_erro)


#------------------------------------------------------------------------------------------------------------------------#

@app.route('/')
def index():
    return render_template('importação-produtos.html')

@app.route('/processar', methods=['POST'])
def processar():
    if request.method == 'POST':
        caminho_arquivo = request.form['arquivo']
        pasta_relatorios = request.form['relatorios']
        
        try:
            arquivo_carregado = pd.read_excel(caminho_arquivo)
        
            analisar_identificacao(arquivo_carregado, "Relatorio_de_Acertos_identificacao", pasta_relatorios)
            analisar_codigo(arquivo_carregado, "Relatorio_de_Acertos_codigo", pasta_relatorios)
            substituir_caractere_excel(caminho_arquivo, ";")
            analisar_referencia(arquivo_carregado, "Relatorio_de_Acertos_Referencia", pasta_relatorios)
        
            analisar_codigo_ean(arquivo_carregado, "Relatorio_de_Acertos_codigo_ean", pasta_relatorios)
        
            analisar_nome(arquivo_carregado,"Relatorio_de_Acertos_Nome", pasta_relatorios)
        
            Tipo(arquivo_carregado, "Relatorio_de_Acertos_Tipo", pasta_relatorios)
        
            Codigo_do_Fornecedor(arquivo_carregado, "Relatorio_de_Acertos_Código do Fornecedor", pasta_relatorios)
        
            Lucro(arquivo_carregado, "Relatorio_de_Acertos_Lucro", pasta_relatorios)
        
            Preco(arquivo_carregado, "Relatorio_de_Acertos_Preço", pasta_relatorios)
        
            unidade_de_medida(arquivo_carregado, "Relatorio_de_Acertos_Unidade_de_Medida", pasta_relatorios)
        
            Peso(arquivo_carregado, "Relatorio_de_Acertos_Peso", pasta_relatorios)
        
            Numero_de_serie(arquivo_carregado, "Relatorio_de_Acertos_Numero_de_serie", pasta_relatorios)
        
            Tributação_ICMS(arquivo_carregado, "Relatorio_de_Acertos_Peso", pasta_relatorios)
        
            IPI(arquivo_carregado, "Relatorio_de_Acertos_Peso", pasta_relatorios)
        
            situacao_tributaria(arquivo_carregado, "Relatorio_de_Acertos_Tributação_ICMS", pasta_relatorios)
        
            Custo(arquivo_carregado, "Relatorio_de_Acertos_Custo", pasta_relatorios)
        
            IAT(arquivo_carregado, "Relatorio_de_Acertos_IAT", pasta_relatorios)
        
            IPPT(arquivo_carregado, "Relatorio_de_Acertos_IPPT", pasta_relatorios)
        
            Origem(arquivo_carregado, "Relatorio_de_Acertos_Origem", pasta_relatorios)
        
            Grupo(arquivo_carregado, "Relatorio_de_Acertos_Grupo", pasta_relatorios)
        
            Fornecedor(arquivo_carregado, "Relatorio_de_Acertos_Fornecedor", pasta_relatorios)
        
            Caminho_da_imagem(arquivo_carregado, "Relatorio_de_Acertos_Caminho_da_imagem", pasta_relatorios)
        
            ICMS(arquivo_carregado, "Relatorio_de_Acertos_ICMS", pasta_relatorios)
        
            Tributacao_especial(arquivo_carregado, "Relatorio_de_Acertos_Tributacao_especial", pasta_relatorios)
        
            Casas_decimais_da_unidade_de_medida(arquivo_carregado, "Relatorio_de_Acertos_Casas_decimais_da_unidade_de_medida", pasta_relatorios)
        
            codigo_de_grupo(arquivo_carregado, "Relatorio_de_Acertos_codigo_de_grupo", pasta_relatorios)
        
            Pesavel(arquivo_carregado, "Relatorio_de_Acertos_codigo_de_Pesavel", pasta_relatorios)
        
            Tipo_de_produto(arquivo_carregado, "Relatorio_de_Acertos_codigo_de_Tipo_produto", pasta_relatorios)
        
            OBS(arquivo_carregado, "Relatorio_de_Acertos_codigo_de_OBS", pasta_relatorios)
        
            Pautas_de_preco(arquivo_carregado, "Relatorio_de_Acertos_codigo_de_Pautas_De_Preco", pasta_relatorios)
            NCM(arquivo_carregado, "Relatorio_de_Acertos_codigo_de_NCM", pasta_relatorios)
            Tributacao_do_simples_nacional_nfe(arquivo_carregado, "Relatorio_de_Acertos_codigo_de_Tributacao_do_simples_nacional_nfe", pasta_relatorios)
            CST_Pis_Cofins_saida(arquivo_carregado, "Relatorio_de_Acertos_codigo_de_CST_Pis_Cofins_saida", pasta_relatorios)
            aliquota_pis_saida(arquivo_carregado, "Relatorio_de_Acertos_aliquota_pis_saida", pasta_relatorios)
            Aliquota_confis_saida(arquivo_carregado, "Relatorio_de_Acertos_Aliquota_confis_saida", pasta_relatorios)
            CST_Pis_Cofins_entrada(arquivo_carregado, "Relatorio_de_Acertos_CST_Pis_Cofins_entrada", pasta_relatorios)
            aliquota_pis_entrada(arquivo_carregado, "Relatorio_de_Acertos_aliquota_pis_entrada", pasta_relatorios)
            aliquota_cofins_entrada(arquivo_carregado, "Relatorio_de_Acertos_aliquota_cofins_entrada", pasta_relatorios)
            Permite_informar_dimensoes(arquivo_carregado, "Relatorio_de_Acertos_Permite_informar_dimensoes", pasta_relatorios)
            CFOP_interna_de_entrada(arquivo_carregado, "Relatorio_de_Acertos_CFOP_interna_de_entrada", pasta_relatorios)
            CFOP_interna_de_saida(arquivo_carregado, "Relatorio_de_Acertos_CFOP_interna_de_saida", pasta_relatorios)
            CFOP_externa_de_entrada(arquivo_carregado, "Relatorio_de_Acertos_CFOP_externa_de_entrada", pasta_relatorios)
            CFOP_externa_de_saida(arquivo_carregado, "Relatorio_de_Acertos_CFOP_externa_de_saida", pasta_relatorios)
            CFOP_interna_de_entrada_devolucao(arquivo_carregado, "Relatorio_de_Acertos_CFOP_interna_de_entrada_devolucao", pasta_relatorios)
            CFOP_interna_de_saida_devolucao(arquivo_carregado, "Relatorio_de_Acertos_CFOP_interna_de_saida_devolucao", pasta_relatorios)
            CFOP_externa_de_entrada_devolucao(arquivo_carregado, "Relatorio_de_Acertos_ CFOP_externa_de_entrada_devolucao", pasta_relatorios)
            CFOP_externa_de_saida_devolucao(arquivo_carregado, "Relatorio_de_Acertos_CFOP_externa_de_saida_devolucao", pasta_relatorios)
            CFOP_interna_de_entrada_transferencia(arquivo_carregado, "Relatorio_de_Acertos_CFOP_interna_de_entrada_transferencia", pasta_relatorios)
            CFOP_interna_de_saida_transferencia(arquivo_carregado, "Relatorio_de_Acertos_CFOP_interna_de_saida_transferencia", pasta_relatorios)
            CFOP_externa_de_entrada_transferencia(arquivo_carregado, "Relatorio_de_Acertos_CFOP_externa_de_entrada_transferencia", pasta_relatorios)
            CFOP_externa_de_saida_transferencia(arquivo_carregado, "Relatorio_de_Acertos_CFOP_externa_de_saida_transferencia", pasta_relatorios)
            Informacao_adicional(arquivo_carregado, "Relatorio_de_Acertos_Informacao_adicional", pasta_relatorios)
            informacao_extra(arquivo_carregado, "Relatorio_de_Acertos_informacao_extra", pasta_relatorios)
            cest(arquivo_carregado, "Relatorio_de_Acertos_cest", pasta_relatorios)
            Tributacao_do_Simples_Nacional_NFCe_ou_SAT(arquivo_carregado, "Relatorio_de_Acertos_Tributacao_do_Simples_Nacional_NFCe_ou_SAT", pasta_relatorios)
            Custo_medio_inicial(arquivo_carregado, "Relatorio_de_Acertos_ Custo_medio_inicial", pasta_relatorios)
            Codigo_da_Lei_complementar(arquivo_carregado, "Relatorio_de_Acertos_Codigo_da_Lei_complementar", pasta_relatorios)
            Indicador_da_exigibilidade_do_ISS(arquivo_carregado, "Relatorio_de_Acertos_ Indicador_da_exigibilidade_do_ISS", pasta_relatorios)

            return "Dados recebidos com sucesso!"# Chame as outras funções de análise aqui...
        except Exception as e:
            mensagem_erro = f"Erro ao processar o arquivo: {str(e)}"
            escrever_pdf(os.path.join(pasta_relatorios, 'Erro_ao_Processar_Arquivo.pdf'), mensagem_erro)
        
    return "Acesso inválido à página de processamento."

if __name__ == '__main__':
    app.run(debug=True)

























































#  Querido Programador

# só deus e eu sei mecher no codigo 
# agora só deus sabe 
# não pergunte, e seja feliz com essa bomba 